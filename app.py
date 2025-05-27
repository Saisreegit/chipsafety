from flask import Flask, request, send_file, jsonify, render_template, url_for, session, redirect
from flask_cors import CORS
from werkzeug.security import check_password_hash
import pandas as pd
import os
import tempfile
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from db import insert_excel_data  # new
from db import get_db_connection


app = Flask(__name__)
CORS(app)

app.secret_key = os.urandom(24)  # Secret key for session


# Dummy credentials (you can later link this to your DB)
USERNAME = 'admin'
PASSWORD = 'password123'

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        print("User fetched from DB:", user)  # 🔍 Debug line

        if user and check_password_hash(user['password'], password):
            session['username'] = user['username']
            return redirect(url_for('dashboard'))  # 👈 updated redirect
        else:
            error = "Invalid username or password"

    return render_template('login.html', error=error)

@app.route('/crms', methods=['GET', 'POST'])
def crms():
    message = ""
    
    work_products = [
        {"name": "Change Request ID", "input_type": "text"},
        {"name": "Start Date", "input_type": "date"},
        {"name": "Phase", "input_type": "select", "options": ["Initiation", "Planning", "Execution", "Monitoring", "Closure"]},
        {"name": "End Date", "input_type": "date"},
        {"name": "Summary of Change", "input_type": "textarea"},
    ]

    if request.method == 'POST':
        action = request.form.get('action')
        submitted_data = {}
        for i, wp in enumerate(work_products, start=1):
            value = request.form.get(f'value{i}')
            submitted_data[wp['name']] = value

        print("Submitted:", submitted_data)

        if action == "Save":
            message = "Page saved successfully!"
        elif action == "Next":
            return redirect(url_for('next_step'))

    return render_template("crms_index.html", work_products=work_products, message=message)

@app.route("/next")
def next_step():
    return "<h2>Welcome to the Next Step</h2><p>You have been redirected successfully.</p>"

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

UPLOAD_FOLDER = tempfile.gettempdir()
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

excel_data = {}  # Cache for uploaded Excel files

@app.route("/")
def index():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template("index.html")

@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('dashboard.html')

@app.route('/work-products')
def work_products():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('work_products.html')

@app.route("/upload", methods=["POST"])
def upload():
    file = request.files["file"]
    if file and file.filename.endswith((".xlsx", ".xls")):
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
        file.save(filepath)
        excel_data[file.filename] = {"path": filepath}
        xl = pd.ExcelFile(filepath)
        return jsonify({"message": "Uploaded", "filename": file.filename, "sheets": xl.sheet_names})
    return jsonify({"error": "Invalid file format"}), 400

@app.route("/edit", methods=["GET"])
def edit():
    filename = request.args.get("filename")
    sheet = request.args.get("sheet")
    file_info = excel_data.get(filename)

    if not file_info or not os.path.exists(file_info["path"]):
        return jsonify({"error": "File not found"}), 404

    df = pd.read_excel(file_info["path"], sheet_name=sheet, dtype=str).fillna("")
    wb = load_workbook(file_info["path"], data_only=True)
    ws = wb[sheet]

    dropdowns = {}
    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            if dv.type == "list" and dv.formula1:
                options = []
                if dv.formula1.startswith("="):
                    try:
                        ref = dv.formula1.strip("=").replace('$', '')
                        if "!" in ref:
                            sheetname, ref = ref.split("!")
                            target_ws = wb[sheetname]
                        else:
                            target_ws = ws
                        min_col, min_row, max_col, max_row = range_boundaries(ref)
                        for row in target_ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                            for cell in row:
                                if cell.value:
                                    options.append(str(cell.value))
                    except Exception:
                        options = []
                else:
                    options = dv.formula1.strip('"').split(",")

                for cell_range in dv.sqref.ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))
                    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                        for cell in row:
                            dropdowns[cell.coordinate] = options

    return jsonify({
        "columns": df.columns.tolist(),
        "data": df.to_dict(orient="records"),
        "dropdowns": dropdowns
    })

@app.route("/save", methods=["POST"])
def save():
    data = request.json
    filename = data.get("filename")
    sheet = data.get("sheet")
    edited_data = data.get("data")

    if filename not in excel_data:
        return jsonify({"error": "File not found"}), 404

    filepath = excel_data[filename]["path"]
    df = pd.DataFrame(edited_data)

    wb = load_workbook(filepath)
    ws = wb[sheet]

    # Clear existing data rows (except header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    for r, row_data in enumerate(df.values, start=2):
        for c, value in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=value)

    wb.save(filepath)

    # Save second column (edited values) to DB
    try:
        insert_excel_data(df)
    except Exception as e:
        return jsonify({"error": f"DB save failed: {str(e)}"}), 500

    return jsonify({"message": "Saved successfully and logged to DB"})

@app.route("/download", methods=["GET"])
def download():
    filename = request.args.get("filename")
    custom_name = request.args.get("custom_name", "Edited_File.xlsx")
    if filename in excel_data:
        filepath = excel_data[filename]["path"]
        return send_file(filepath, as_attachment=True, download_name=custom_name)
    return jsonify({"error": "File not found"}), 404

if __name__ == "__main__":
    app.run(debug=True)
